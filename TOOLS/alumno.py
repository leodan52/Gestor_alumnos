# Clases para alumno

from TOOLS.tools import *
import json, csv
from tabulate import tabulate as tabla
import TOOLS.busqueda as bq
#from shutil import rmtree
from datetime import datetime
import re
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font

#*-------------------------------------------------------------------------------------------*
#*------ Clases -----------------------------------------------------------------------------*
#*-------------------------------------------------------------------------------------------*


class TodosMisAlumnos:
	version = '1.0.0'

	def __init__(self,lista_,ruta_pdf, ruta_base_datos):

		''' Clase que contiene las listas de todos los alumnos '''
		self.__version_clase = TodosMisAlumnos.version
		self.__lista = lista_
		self.ruta_pdf = ruta_pdf
		self.ruta_base_datos = ruta_base_datos
		self.ruta_h = os.path.join(os.path.dirname(self.ruta_base_datos), ".CACHE")
		self.ruta_historial = f'historial_cambios.txt'
#		self.base_status = f'base_estatus.json'
#		self.lineasF_status = f'lineas_estatus.json'
		self.Arch_dir_historial()
		self.Data = dict()
		self.estructuraArbol = dict()
		self.CambiosSinGuardar = dict()
		self.ClavesBusqueda = []

		self.Control_version()
		self.EntregarPDF()
		self.UsedData()

	def __getitem__(self, index):
		return self.__lista[index]

	def __iter__(self):
		yield from self.__lista

	@property
	def version_clase(self):
		return self.__version_clase


	#------- Exportar los alumnos desde la base de datos-------------------------------------------
	@classmethod
	def Extraer_instancia(cls, ruta_Base, ruta_PDF):

		''' Extraer los datos de los alumnos  de un archivo existente '''

		try:
			instance_ = CargarBinario(ruta_Base)
			if hasattr(instance_, 'version'):
				version_instance = instance_.version_clase
				version_actual = cls.version
			else:
				raise AttributeError('La instancia guardada no tiene version')

			print(version_instance, version_actual)

			if version_instance != version_actual:
				raise ValueError(f'Version de instancia ({version_instance}) no coicide con version actual ({version_actual}) de la clase')

			return instance_

		except (FileNotFoundError, AttributeError, ValueError) as e:
			print(f'Se desplegó la siguiente Excepcion:')
			print(f'\t{type(e).__name__}: {e}')
			print('Se procede con una instancia vacia nueva')
			os.makedirs(os.path.split(ruta_Base)[0], exist_ok = True)
			return cls([], ruta_PDF, os.path.split(ruta_Base)[0])

	#---------------------------------------------------------------------------------------------

	def Arch_dir_historial(self):

		''' Construye las rutas de archivos de historial '''

		self.ruta_historial = os.path.join(self.ruta_h, self.ruta_historial)
#		self.base_status = os.path.join(self.ruta_h, self.base_status)
#		self.lineasF_status = os.path.join(self.ruta_h, self.lineasF_status)

	#---------------------------------------------------------------------------------------------

	def Control_version(self):

		''' Guarda hasta 50 versiones pasadas en formato JSON '''

		versiones_ruta = os.path.join(self.ruta_h, "Historial_versiones")
		maxi = 50

		os.makedirs(versiones_ruta, exist_ok = True)

		versiones = os.listdir(versiones_ruta)

		date = datetime.today().strftime('%Y-%m-%d %H_%M_%S')
		add_version = os.path.join(versiones_ruta, f'{date}.json')

		if versiones == []:
			self.Todos2JSON(add_version)
			return

		versiones.sort()
		ruta_ulti = os.path.join(versiones_ruta, versiones[-1])
		ulti = self.ImportJSON(ruta_ulti, False)

		nueva = self.Todos2JSON("", False)

		if nueva == ulti:
			return

		self.Todos2JSON(add_version)
		versiones.append("")

		if len(versiones) > maxi:
			n = len(versiones)
			delta = n - maxi
		else:
			return

		for i in versiones:
			if versiones.index(i) == delta:
				break
			os.remove(os.path.join(versiones, i))

	#--------------------------------------------------------------------------------------------

	def EntregarPDF(self):

		''' Revisa la base de PDFs para verificar si cada alumno entregó o no su archivo PDF '''

		try:
			cursos = os.listdir(self.ruta_pdf)
		except FileNotFoundError:
			os.mkdir(self.ruta_pdf)
			cursos = os.listdir(self.ruta_pdf)

		entregado = dict()

		for curso in cursos:

			try:
				PDFs = os.listdir( os.path.join(self.ruta_pdf, curso))
			except NotADirectoryError:
				continue

			#datos_curso = curso.split("_")

			for PDF in PDFs:

				rutaPDF = os.path.join(self.ruta_pdf, curso, PDF)
				PDF = quitaracentos(PDF)

				if "pdf" not in PDF.split("."):
					continue

				entregado[PDF.replace(".pdf","")] = rutaPDF

		for alumno in self.__lista:
			nombre = alumno.comparar_nombre

			if nombre in entregado:
				alumno.NR_entregado = "Sí"
				alumno.RutaPDF = entregado[nombre]
			else:
				alumno.NR_entregado = "*No*"

	#---------------------------------------------------------------------------------------------

	def UsedData(self):

		''' Mantiene un control de los datos en común para cada alumno '''

		self.Data["Plantel"] = []
		self.Data["Curso"] = []
		self.Data["Horario"] = []

		self.Data["Carrera"] = []
		self.Data["Centro Universitario"] = []

		for alumno in self.__lista:
			if alumno.plantel not in self.Data["Plantel"]:
				self.Data["Plantel"].append(alumno.plantel)
			if alumno.curso not in self.Data["Curso"]:
				self.Data["Curso"].append(alumno.curso)
			if alumno.horario not in self.Data["Horario"]:
				self.Data["Horario"].append(alumno.horario)

			if alumno.carrera not in self.Data["Carrera"]:
				self.Data["Carrera"].append(alumno.carrera)
			if alumno.CU not in self.Data["Centro Universitario"]:
				self.Data["Centro Universitario"].append(alumno.CU)

			claves = procesar(alumno.nombre).lower().split(" ")
			for clave in claves:
				if clave not in self.ClavesBusqueda:
					self.ClavesBusqueda.append(clave)

		for i in self.Data:
			self.Data[i].sort()


	#---------------------------------------------------------------------------------------------

	def Cambios_realizados(self,cadena):

		''' Hace un seguimiento de los cambios sin guardar realizados a la base de datos '''

		fecha = datetime.today().strftime('%Y-%m-%d %H:%M:')

		try:
			self.CambiosSinGuardar[fecha].append(cadena)
		except KeyError:
			self.CambiosSinGuardar[fecha] = [cadena]
		except TypeError:
			self.CambiosSinGuardar = {fecha: [cadena]}


	def Eventos_control(self,cadena):

		''' Hace seguimiento de eventos relevantes para la gestión '''

		Agregar2TXT(self.ruta_historial, [cadena])


	#---------------------------------------------------------------------------------------------

	def add_cHistorial(self):

		''' Al importar los cambios a la base de datos, guarda los cambios en el archivo historial y
			vacía el seguimiento de cambios sin guardar '''

#		if self.CambiosSinGuardar != []:
#			Agregar2TXT(self.ruta_historial, self.CambiosSinGuardar)
#			self.CambiosSinGuardar = []

		if len(self.CambiosSinGuardar) == 0:
			return

		lista = []

		for i in self.CambiosSinGuardar:
			if i not in lista:
				lista.append(i)
			lista += self.CambiosSinGuardar[i]

		Agregar2TXT(self.ruta_historial, lista)
		self.CambiosSinGuardar = dict()


	#--------------------------------------------------------------------------------------------

	def AddAlumno(self,lista):

		''' Agrega alumno a la lista de alumnos '''

#		n, nr, c, cu, curso, plantel, horario

#		fecha = datetime.today().strftime('%Y-%m-%d %H:%M')

		cambio = (f'\tSe agregó el alumno {lista[0]}:\n\t\tNúmero de registro:'
			+ f' {lista[1]}, Carrera: {lista[2]}, Centro universitario: {lista[3]}, '
			+ f'Curso: {lista[4]}, Plantel: {lista[5]}, Horario: {lista[6]}')

		self.__lista.append(Alumno(*lista))
		self.Cambios_realizados(cambio)
		self.UsedData()

	#-------------------------------------------------------------------------------------------------

	def RmAlumno(self,indice):

		''' Eliminar alumno '''

		d = self.__lista[indice].datos

#		fecha = datetime.today().strftime('%Y-%m-%d %H:%M')
		cambio= f'\tSe eliminó al alumno {d["Nombre"]}:\n\t\t'
		for i in d:
			if i == "Nombre" or i == "¿PDF entregado?":
				continue

			aux = f'{i}: {d[i]}, '
			cambio = cambio + aux

		self.__lista.pop(indice)
		self.Cambios_realizados(cambio)
		self.UsedData()

	#--------------------------------------------------------------------------------------------
	@MensajeClase
	def organizarPDF(self, Consola = True):

		''' Manipula la base (directorio) que contiene los PDFs y los organiza en directorios por salón.
			El PDF debe renombrarse usado el nombre del alumno '''

		try:
			PDFs = []
			for ruta, _, archivos in os.walk(self.ruta_pdf):
				aux = list(map(lambda x: os.path.join(ruta, x), archivos))
				PDFs = PDFs + aux
		except FileNotFoundError:
			os.mkdir(self.ruta_pdf)
			PDFs = []

		control = [datetime.today().strftime('%Y-%m-%d %H:%M'), "\tArchivos desplazados:"]
		PDFs = list(filter(lambda x: os.path.isfile(x) and x.endswith('.pdf'), PDFs))

		for PDF in PDFs:
			pdf_string = PDF2Cadena(PDF)

			alumnos = list(filter(lambda x: TodosMisAlumnos.__buscarString(x.comparar_nombre, pdf_string), self.__lista))

			if len(alumnos) > 1:
				print(f'Warning: Hay {len(alumnos)} alumnos que coinciden con el PDF.')
				print('\t', alumnos)
				print('Se procede con el primero.')
			elif len(alumnos) == 0:
				continue

			alumno = alumnos[0]

			curso = f'{alumno.plantel}_{alumno.curso}_{alumno.horario}'
			nombre_alumno = alumno.comparar_nombre.title() + alumno.bandera + '.pdf'
			nueva_ruta = os.path.join(self.ruta_pdf, curso, nombre_alumno)

			if PDF == nueva_ruta:
				continue

			try:
				os.rename(PDF, nueva_ruta)
			except FileNotFoundError:
				os.mkdir(os.path.dirname(nueva_ruta))
				os.rename(PDF, nueva_ruta)
			except FileExistsError:
				raise IOError("No sé que pasa")

			if Consola:
				print("\nArchivos desplazados:")
				print(f'{PDF} --> {nueva_ruta}/\n')
			else:
				mensaje = f'\t\t{PDF} --> {nueva_ruta}/'
				control.append(mensaje)

		if len(control) > 2:
			mensaje = "\n".join(control)
			self.Eventos_control(mensaje)

		self.EntregarPDF()

	@staticmethod
	def __buscarString(string, pdf_string):
		string_ = string.split(' ')

		check = list(map(lambda x: x in pdf_string, string_))

		return all(check)


	#--------------------------------------------------------------------------------------------
	@MensajeClase
	def EscanearPDFs(self):

		Error = dict()
		indice = 0

		for alumno in self.__lista:
			nuevo = alumno.ValidarPDF()
			if hasattr(alumno, "CoinPorcent"):
				if alumno.CoinPorcent < 50.0:
					Error[alumno.nombre] = (alumno.RutaPDF, "Archivo PDF")
				elif alumno.CoinPorcent == 101:
					print(f'El archivo {alumno.RutaPDF} no se puede leer')
				elif alumno.CoinNum == 0.0:
					Error[alumno.nombre] = (alumno.RutaPDF, "Número de registro")

			if nuevo != None:
				self.MotorEditar_GUI(indice, "Número de registro", nuevo)

			indice += 1

		self.ErrorPDFs = Error

	#--------------------------------------------------------------------------------------------
	@MensajeClase
	def GenerarListas(self, archivo = "Listas completas.txt"):

		''' Genera listas tabuladas de los alumnos divididos por salones '''

		#print("Has invocado al héroe japonés")

		Salones = dict()

		combinaciones = product(self.Data["Plantel"], self.Data["Curso"], self.Data["Horario"])

		for plantel, curso, horario in combinaciones:

			cadena_comparar = f'{plantel}-{curso}-{horario}'

			salon = filter(lambda alumno: cadena_comparar == f'{alumno.plantel}-{alumno.curso}-{alumno.horario}' , self.__lista)
			salon = list(salon)

			if salon == []:
				continue
			else:
				salon.sort()

			Salon_actual = f'\n\n** Lista del plantel *{plantel}*, curso *{curso}* del horario *{horario}* **\n\n'
			Salones[Salon_actual] = []

			for alumno in salon:
				datos_alumno = [alumno.nombre,alumno.nRegistro,alumno.carrera,alumno.CU,alumno.NR_entregado, alumno.admitido]
				Salones[Salon_actual].append(datos_alumno)

		titulos = ["Nombre","Num. de registro", "Carrera", "Centro", "PDF enviado", "¿Admitido?"]

		with open(archivo, "w") as salida:

			for salon in Salones:
				print(salon, file=salida)

				check_dictamen = ChecarDictamen(Salones[salon])

				if not check_dictamen and titulos[-1] == "¿Admitido?":
					titulos.pop(-1)
				elif titulos[-1] != "¿Admitido?":
					titulos.append("¿Admitido?")

				print(tabla(Salones[salon], headers=titulos), file=salida)


			print("\n", file=salida)

	@MensajeClase
	def GenerarListasExcel(self, archivo = "Listas_completas.xlsx"):

		wb = Workbook()
		ws = wb.active
		ft_headers = Font(bold=True)
		n = 0

		combinaciones = product(self.Data["Plantel"], self.Data["Curso"], self.Data["Horario"])
		headers = ["Nombre","Num. de registro", "Carrera", "Centro", "PDF enviado", "¿Admitido?"]

		for plantel, curso, horario in combinaciones:

			cadena_comparar = f'{plantel}-{curso}-{horario}'

			salon = filter(lambda alumno: cadena_comparar == f'{alumno.plantel}-{alumno.curso}-{alumno.horario}', self.__lista)
			salon = list(salon)

			if salon == []:
				continue
			else:
				salon.sort()

			if n == 0:
				ws.title = f'{plantel}{horario}'.replace(' ', '')
			else:
				ws = wb.create_sheet(f'{plantel}{horario}'.replace(' ', ''))

			ws.append([cadena_comparar])
			ws.append(headers)
			for cell in ws[2]:
				cell.font = ft_headers

			for alumno in salon:
				ws.append([alumno.nombre,alumno.nRegistro,alumno.carrera,alumno.CU,alumno.NR_entregado, alumno.admitido])

			n += 1

		wb.save(archivo)
		wb.close()

	#--------------------------------------------------------------------------------------------

	def MotorEditar_GUI(self,indice,dato,nuevo):

		''' Motor para edición de datos de alumnos para GUI'''

		viejo = self.__lista[indice].datos[dato]
#		fecha = datetime.today().strftime('%Y-%m-%d %H:%M')

		cambio = (f'\tCambio en {self.__lista[indice].nombre} del plantel '
			+ f'{self.__lista[indice].plantel}, curso {self.__lista[indice].curso}, '
			+ f'horario {self.__lista[indice].horario}:\n\t\t Cambió {dato}: {viejo} --> {nuevo}')

		self.__lista[indice].ActualizarDato(dato,nuevo)
		self.Cambios_realizados(cambio)
		self.UsedData()


	#--------------------------------------------------------------------------------------------

	def Buscar_(self,cadena):
		indice = bq.buscar_GUI(procesar(cadena),self.__lista)

		return indice

	#--------------------------------------------------------------------------------------------

	def GenerarArbol(self):

		indice = 0
		self.estructuraArbol = dict()

		for i in self.__lista:
			plantel, curso, horario = i.DatosSalon

			if plantel not in self.estructuraArbol:
				self.estructuraArbol[plantel] = dict()

			if curso not in self.estructuraArbol[plantel]:
				self.estructuraArbol[plantel][curso] = dict()

			if horario not in self.estructuraArbol[plantel][curso]:
				self.estructuraArbol[plantel][curso][horario] = []

			self.estructuraArbol[plantel][curso][horario].append(f'{i.nombre}&{indice}')
			indice += 1

	def getArbol(self):
		self.GenerarArbol()

		return self.estructuraArbol

	#--------------------------------------------------------------------------------------------
	@MensajeClase
	def AbrirPDF_GUI(self,indice):

		''' Abrir el documento PDF desde el GUI '''

		ruta = self.__lista[indice].RutaPDF

		if ruta == "Sin ruta":
			return
		abrete(ruta)

	#------------------JSON Exportar -------------------------------------------------------------
	@MensajeClase
	def Todos2JSON(self, archivo = "alumnos.json", salida_file = True):

		''' Exporta los datos de los alumnos en listas a un archivo JSON '''

		todos = dict()
		todos["Alumnos"] = []
		todos["Total alumnos"] = 0
		total = 0
		todos["Planteles"] = []
		todos["Cursos"] = []
		grupos = []
		total_nr = 0

		for i in self.__lista:
			todos["Alumnos"].append(i.datos_json)
			total += 1

			if i.datos["Plantel"] not in todos["Planteles"]:
				todos["Planteles"].append(i.datos["Plantel"])
			if i.datos["Curso"] not in todos["Cursos"]:
				todos["Cursos"].append(i.datos["Curso"])
			if i.datos["Número de registro"] not in ["n/a", "No data"]:
				total_nr += 1

			grupo = i.datos["Plantel"] + i.datos["Curso"] + i.datos["Horario"]

			if grupo not in grupos:
				grupos.append(grupo)

		todos["Total alumnos"] = total
		todos["Total grupos"] = len(grupos)
		try:
			todos["% de registro entregados"] = 100*total_nr/total
		except ZeroDivisionError:
			todos["% de registro entregados"] = "No data"


		cadena = json.dumps(todos, ensure_ascii=False, indent = 4)

		if salida_file:
			with open(archivo, "w") as salida:
				print(cadena, file=salida)
		else:
			return todos

	#-------- Importar JSON -----------------------------------------------------------------
	@MensajeClase
	def ImportJSON(self, archivo = "alumnos.json", importar = True):

		''' Importa los datos de los alumnos contenidos en un JSON a las listas. Se debe actualizar
			la base de datos principal posteriormente '''

		with open(archivo, "r") as entrada:
			lineas = entrada.readlines()
			lineas = "".join(lineas)
			lineas = lineas.replace("\t","")
			lineas = lineas.replace("\n","")

		datos = json.loads(lineas)

		if len(datos) == 0:
			return

		if importar:
			pass
		else:
			return datos

		for alumno in datos["Alumnos"]:
			n, nr, c = alumno["Nombre"],alumno["Número de registro"],alumno["Carrera"]
			cu = alumno["Centro Universitario"]
			curso,plantel,horario = alumno["Curso"],alumno["Plantel"],alumno["Horario"]

#			self.lista.append(Alumno(n,nr,c,cu,curso,plantel,horario))

			argumento = [n, nr, c, cu, curso, plantel, horario]

			self.AddAlumno(argumento)


#		fecha = datetime.today().strftime('%Y-%m-%d %H:%M')
#		cambio = f'{fecha}:\n\tEl archivo {archivo} fue cargado. Mire el contenido para más información'
#		self.Cambios_realizados(cambio)
#		self.UsedData()

	#------------- Importar CSV ---------------------------------------------------------------

	@MensajeClase
	def ImportarCSV(self, archivo, datos_curso):

		''' Importar datos desde un archivo CSV '''

		with open(archivo, newline="") as entrada:
			lector = csv.reader(entrada)

			for row in lector:
				if row == [] or row == [""]:
					continue

				newRow = row + datos_curso
#				self.lista.append(Alumno(*newRow))
				self.AddAlumno(newRow)

#		fecha = datetime.today().strftime('%Y-%m-%d %H:%M')
#		cambio = f'{fecha}:\n\tEl archivo {archivo} fue cargado. Mire el contenido para más información'
#		self.Cambios_realizados(cambio)
#		self.UsedData()

	#------------- Exportar CSV ---------------------------------------------------------------
	@MensajeClase
	def ExportarCSV(self, ruta_archivo, *indices):

		''' Crea un archivo CSV con los alumnos seleccionados '''

		with open(ruta_archivo, "w", newline="") as salida:
			writer = csv.writer(salida)

			for i in indices:
				alumno = self.__lista[i].DatosUDG
				writer.writerow(alumno)

	#----------- Exportar algunos alumnos JSON ------------------------------------------------

	@MensajeClase
	def ExportarAulaJSON(self, ruta_archivo, *indices):

		''' Crea un archivo JSON con los alumnos seleccionados '''

		salida = {"Alumnos" : []}

		for i in indices:
			salida["Alumnos"].append(self.__lista[i].datos_json)

		salida = json.dumps(salida, ensure_ascii=False, indent = 4)

		with open(ruta_archivo, "w") as archivo:
			print(salida, end="", file=archivo)


	#-------- Leer dictamen para conocer admitidos -----------------------------------------
	@MensajeClase
	def LeerDictamen(self, archivo):

		''' Lee el dictamen de admitidos UDG en formato PDF para verificar admitidos '''

		cadena = PDF2Cadena(archivo)

		print(f'Mensaje de la clase "{type(self).__name__}": El archivo dictamen tiene una longitud de {len(cadena)}')

		self.num_Admitidos = 0

		for alumno in self.__lista:
			alumno.ConfirmarAdmision(cadena)

			if alumno.admitido == "*Sí*":
				self.num_Admitidos += 1



#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------

class Alumno:

	version = '1.0.0'

	def __init__(self, nombre, numRegistro, carrera, CU, curso, plantel, horario):

		''' Clase alumnos. Contiene la información de un alumnos '''

		# nombre: Nombre del alumno
		# numRegistro: Proporcionado al hacer tramites UDG
		# carrera: Carrera a aspirar
		# CU: Centro universitario de la red
		# Curso: Nombre del curso en Lumiere
		# plantel: Plantel de Lumiere donde asiste el alumno
		# horario: Horario en el que el alumno asiste

		self._nombre = QuitarMEspacios(nombre)
		self._nRegistro = numRegistro
		self._carrera = QuitarMEspacios(carrera)
		self._CU = CU
		self._curso = curso
		self._plantel = plantel
		self._horario = horario
		self.NR_entregado = "No data"
		self.bandera = ""
		self.RutaPDF = "Sin ruta"
		self.admitido = "No data"

		self.RellenarValoresVacios()

	@property
	def nombre(self):
		return self._nombre

	@property
	def nRegistro(self):
		return self._nRegistro

	@property
	def carrera(self):
		return self._carrera

	@property
	def CU(self):
		return self._CU

	@property
	def curso(self):
		return self._curso

	@property
	def plantel(self):
		return self._plantel

	@property
	def horario(self):
		return self._horario

	#--------------------------------------------------------------------------------------------

	def __repr__(self):
		representacion = f'< Objeto : Alumno, Nombre: \'{self._nombre}\', Carrera: \'{self._carrera}\', ' + \
			f'CU: \'{self._CU}\', Lumiere: \'{self._plantel}\' - \'{self._horario}\' - \'{self._curso}\'> '

		return representacion

	#--------------------------------------------------------------------------------------------

	def __lt__(self, other):
		return self.comparar_nombre < other.comparar_nombre

	#--------------------------------------------------------------------------------------------

	def RellenarValoresVacios(self):

		aux = [self._nRegistro, self._carrera, self._CU]

		while "" in aux:
			indice = aux.index("")
			if indice == 0:
				self._nRegistro = "n/a"
			elif indice == 1:
				self._carrera = "n/a"
			else:
				self._CU = "n/a"

			aux = [self._nRegistro, self._carrera, self._CU]


	#--------------------------------------------------------------------------------------------

	def ReiniciarInfoPDFs(self):
		self.NR_entregado = "No data"
		self.RutaPDF = "Sin ruta"

	#--------------------------------------------------------------------------------------------

	def ActualizarDato(self,dato,nuevo):

		''' Cambiar determinado dato por el alumnos. Si el dato no es un atributo, retornará 0,
			caso contrario, 1 '''

		if dato == "Número de registro":
			self._nRegistro = nuevo
		elif dato == "Carrera":
			self._carrera = nuevo
		elif dato == "Centro Universitario":
			self._CU = nuevo
		elif dato == "Nombre":
			self._nombre = nuevo
		elif dato == "Curso":
			self._curso = nuevo
		elif dato == "Plantel":
			self._plantel = nuevo
		elif dato == "Horario":
			self._horario = nuevo
		else:
			return 0

		return 1
	#--------------------------------------------------------------------------------------------

	def ValidarPDF(self):
		if self.RutaPDF == "Sin ruta":
			return
		try:
			cadena = PDF2Cadena(self.RutaPDF)
		except fitz.fitz.EmptyFileError:
			print("\nERROR fitz.fitz.EmptyFileError")
			print(f'[self.RutaPDF] no es un archivo PDF\n')
			return

		if cadena == "":
			self.CoinPorcent = 101
		else:
			self.CoinPorcent = BuscarEnCadena(self.comparar_nombre, cadena)

		if self._nRegistro != "n/a":
			self.CoinNum = BuscarEnCadena(self._nRegistro, cadena)
		else:
			self.CoinNum = -1
			return self.ExtraerNumRegistro(cadena)


	def ExtraerNumRegistro(self, cadena):

		encontrado = re.findall(r'\D\d{7}\D', str(cadena))

		sin_repetidos = []

		for clave in encontrado:
			if clave[1:-1] not in sin_repetidos:
				sin_repetidos.append(clave[1:-1])

		if len(sin_repetidos) == 1:
			return sin_repetidos[0]


	#--------------------------------------------------------------------------------------------

	def ConfirmarAdmision(self, dictamen):
		if self._nRegistro == "n/a":
			return

		porc_coincidencia = BuscarEnCadena(self._nRegistro, dictamen)

		if porc_coincidencia == 100:
			self.admitido = "*Sí*"
		else:
			self.admitido = " No"

	#--------------------------------------------------------------------------------------------

	@property
	def comparar_nombre(self):
		return quitaracentos(self._nombre)

	#--------------------------------------------------------------------------------------------

	@property
	def DatosSalon(self):
		salida = ([
			self._plantel,
			self._curso,
			self._horario
			])
		return salida

	#--------------------------------------------------------------------------------------------

	@property
	def DatosUDG(self):
		salida = ([
			self._nombre,
			self._nRegistro,
			self._carrera,
			self._CU
		])

		return salida

	#--------------------------------------------------------------------------------------------
	@property
	def datos(self):

		''' Genera diccionario que contiene la información del alumno '''

		datos_ = ({ "Nombre" : self._nombre, "Número de registro" : self._nRegistro,
					"Carrera" : self._carrera, "Centro Universitario" : self._CU,
					"Curso" : self._curso, "Plantel" : self._plantel, "Horario" : self._horario,
					"¿PDF entregado?" : self.NR_entregado, "¿Admitido?" : self.admitido})

		if self.admitido == "No data":
			datos_.pop("¿Admitido?")

		return datos_

	#-------------------------------------------------------------------------------------------
	@property
	def datos_json(self):

		''' Genera diccionario con los datos del alumno como preparación para exportar a JSON '''

		datos = ({	"Nombre" : self._nombre, "Número de registro" : self._nRegistro,
					"Carrera" : self._carrera, "Centro Universitario" : self._CU,
					"Curso" : self._curso, "Plantel" : self._plantel, "Horario" : self._horario,
					"¿Admitido?" : self.admitido})

		return datos
